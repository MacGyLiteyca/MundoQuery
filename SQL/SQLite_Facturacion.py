import os
import sqlite3
from datetime import datetime

class DatabaseManager:
    def __init__(self, db_file):
        self.conn = sqlite3.connect(db_file)
        self.cursor = self.conn.cursor()

    def create_tables(self):
        # Crear las tablas si no existen
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS archivos (
                NombreArchivo TEXT PRIMARY KEY,
                FechaCreacion TEXT,
                FechaModificacion TEXT,
                Estado TEXT,
                FechaProceso TEXT
            )
        ''')
        # ... (crear otras tablas si es necesario)

    def insert_file_record(self, file_name, creation_date, modification_date):
        self.cursor.execute("INSERT OR IGNORE INTO archivos VALUES (?, ?, ?, 'Sin procesar', NULL)",
                           (file_name, creation_date, modification_date))

    def update_file_status(self, file_name, status, process_date):
        self.cursor.execute("UPDATE archivos SET Estado=?, FechaProceso=? WHERE NombreArchivo=?",
                           (status, process_date, file_name))

    def get_files_to_process(self):
        self.cursor.execute("SELECT NombreArchivo FROM archivos WHERE Estado='Sin procesar'")
        return [row[0] for row in self.cursor.fetchall()]

    def insert_data_into_qryas(self, file_name, data):
        # Aquí puedes personalizar la inserción según la estructura de QRYAS
        for row in data:
            self.cursor.execute("INSERT INTO QRYAS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", row)

    def insert_data_into_facturado(self, data):
        for row in data:
            # Ensure the row has the correct number of elements
            if len(row) != 48:  # Adjust if the number of columns in Facturado changes
                print(f"Error: Row has incorrect number of elements: {len(row)}")
                continue

            # Create the SQL query with the correct number of placeholders
            sql = "INSERT INTO Facturado VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"

            # Execute the query
            self.cursor.execute(sql, row)

    def commit(self):
        self.conn.commit()

    def close(self):
        self.conn.close()


import os
import openpyxl

class ExcelManager:
  def __init__(self, directory):
    self.directory = directory

  def get_excel_files(self):
        excel_files = []
        for root, _, files in os.walk(self.directory):
            for file in files:
                if file.endswith('.xlsx'):
                    excel_files.append(os.path.join(root, file))
        return excel_files

  def read_sheets(self, file_path, sheet_names=None):
    data = {}
    try:
      workbook = openpyxl.load_workbook(filename=file_path, read_only=True)  # Open in read-only mode

      # Read all sheets if no sheet names specified
      if not sheet_names:
        sheet_names = workbook.sheetnames

      for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
          print(f"Warning: Sheet '{sheet_name}' not found in '{file_path}'.")
          continue
        sheet = workbook[sheet_name]
        data[sheet_name] = []
        for row in sheet.iter_rows(min_row=2):  # Skip header row
          data[sheet_name].append([cell.value for cell in row])  # Extract cell values
    except Exception as e:
      print(f"Error reading Excel file '{file_path}': {str(e)}")
    return data




if __name__ == "__main__":
    db = DatabaseManager('D:\SQLite\Facturacion_Claro.db')
    db.create_tables()

    file_manager = ExcelManager("C:\\Users\\User\\OneDrive - LITEYCA DE COLOMBIA S.A.S\\Compartido\\ProyectoClaroAnt\\Claro Bot\\Facturacion")

    files = file_manager.get_excel_files()
    
    # Imprimir la cantidad de archivos encontrados
    print(f"Se encontraron {len(files)} archivos.")

    for file in files:
        file_name = os.path.basename(file).split('.')[0]
        creation_date = datetime.fromtimestamp(os.path.getctime(file)).strftime('%Y-%m-%d %H:%M:%S')
        modification_date = datetime.fromtimestamp(os.path.getmtime(file)).strftime('%Y-%m-%d %H:%M:%S')

        db.insert_file_record(file_name, creation_date, modification_date)

        files_to_process = db.get_files_to_process()
        for file_to_process in files_to_process:
            data = file_manager.read_sheets(os.path.join(file_manager.directory, file_to_process + '.xlsx'), ['FINAL DE FACT'])
            db.insert_data_into_facturado(file_to_process, data)
            db.update_file_status(file_to_process, 'Procesado', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        db.commit()

    db.close()